import sys
import os
from collections import defaultdict

import ssg.components

from .most_used_rules import _sorted_dict_by_num_value
from .common import generate_output

PYTHON_2 = sys.version_info[0] < 3

if not PYTHON_2:
    from .most_used_rules import _get_profiles_for_product
    from ..controleval import (
        load_controls_manager,
        load_product_yaml,
    )


def _list_rule_components(components, rules_list, components_out):
    for component_name, component in components.items():
        if len(set(component.rules).intersection(set(rules_list))) > 0:
            components_out[component_name].update(set(component.rules).intersection(set(rules_list)))


def load_components(product):
    product_yaml = load_product_yaml(product)
    product_dir = product_yaml.get("product_dir")
    components_root = product_yaml.get("components_root")
    if components_root is None:
        return None
    components_dir = os.path.abspath(os.path.join(product_dir, components_root))
    return ssg.components.load(components_dir)


def _process_all_products_from_controls(components_out, products):
    if PYTHON_2:
        raise Exception("This feature is not supported for python2.")

    for product in products:
        components = load_components(product)
        if components is None:
            continue
        controls_manager = load_controls_manager("./controls/", product)
        for profile in _get_profiles_for_product(controls_manager, product):
            _list_rule_components(components, profile.rules, components_out)


def command_used_rules_per_component(args):
    components = defaultdict(set)

    _process_all_products_from_controls(components, args.products)

    sorted_components = _sorted_dict_by_num_value(components)
    # csv_header = "component_name,count_of_profiles"
    # generate_output(sorted_components, args.format, csv_header)

    from openpyxl import Workbook
    from openpyxl.styles import Font


    def create_spreadsheet(entries):
        # Create a new Workbook
        wb = Workbook()
        
        # Remove default sheet created by openpyxl
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Set font style for hyperlink
        font = Font(color="0563C1", underline="single")
        
        sorted_keys = sorted(components.keys())

        for entry_name in sorted_keys:
            items = components[entry_name]
            # print(entry_name)

            # if len(items) < 20:
            #     continue

            # Create a new worksheet
            ws = wb.create_sheet(title=entry_name)
            
            # Write items to the worksheet
            for i, item in enumerate(items, start=0):
                            # Write the item
                ws.cell(row=i+1, column=1, value=item)
                
                # Create hyperlink
                url = f"https://complianceascode.github.io/content-pages/components/{list(args.products)[0]}/{entry_name}.html#{item}"
                cell = ws.cell(row=i+1, column=1)
                cell.hyperlink = url
                cell.value = f'=HYPERLINK("{url}", "{item}")'
                cell.hyperlink.target_mode = "External"
                cell.font = font  # Set font style for hyperlink
        
        # Save the workbook
        wb.save("entries_spreadsheet.xlsx")

    # Example entries

    # Create the spreadsheet
    create_spreadsheet(components)
